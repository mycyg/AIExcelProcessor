import asyncio
import json
import multiprocessing
import os
import dataclasses
import math
import tempfile
from typing import Any, Dict, Optional, List
import httpx
import pandas as pd
import openpyxl # Using openpyxl for robust, low-memory streaming read
from volcenginesdkarkruntime import AsyncArk

from config import ProcessingConfig

def process_entrypoint(
    worker_id: int,
    num_workers: int,
    max_concurrency: int,
    api_key: Optional[str],
    input_file: str,
    output_file: str,
    config_dict: Dict[str, Any],
    progress_queue: "multiprocessing.Queue",
):
    """
    This is the entrypoint for each worker process.
    It reads a chunk of the input file, processes it concurrently, and writes to an output file.
    This approach avoids loading all tasks into memory at once.
    """
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop=loop)
    
    http_client = httpx.AsyncClient(
        limits=httpx.Limits(max_connections=max_concurrency, max_keepalive_connections=max_concurrency),
        timeout=httpx.Timeout(config_dict.get("api_timeout", 180))
    )
    client = AsyncArk(api_key=api_key, http_client=http_client)

    async def process_chunk(chunk: List[Dict[str, Any]], f_out: Any, sem: asyncio.Semaphore):
        tasks = []
        for record_data in chunk:
            await sem.acquire()
            tasks.append(
                loop.create_task(
                    worker(client, record_data, f_out, sem, progress_queue, config_dict)
                )
            )
        if tasks:
            await asyncio.gather(*tasks)

    async def inner():
        sem = asyncio.Semaphore(max_concurrency)
        chunk_size = max_concurrency * 2 # A reasonable chunk size
        
        with open(input_file, 'r', encoding='utf-8') as f_in, \
             open(output_file, 'w', encoding='utf-8') as f_out:
            
            chunk = []
            for i, line in enumerate(f_in):
                if i % num_workers != worker_id:
                    continue
                
                chunk.append(json.loads(line))
                
                if len(chunk) >= chunk_size:
                    await process_chunk(chunk, f_out, sem)
                    chunk = []
            
            # Process the final remaining chunk
            if chunk:
                await process_chunk(chunk, f_out, sem)

        await http_client.aclose()

    try:
        loop.run_until_complete(inner())
    finally:
        progress_queue.put(None) # Signal that this worker is done

async def worker(
    client: AsyncArk,
    record_data: Dict[str, Any],
    f_out: Any,
    sem: asyncio.Semaphore,
    progress_queue: "multiprocessing.Queue",
    config_dict: Dict[str, Any],
):
    """
    This final version correctly implements the automatic prompt generation
    logic that was missing from the SDK mode, making it consistent with the
    non-SDK mode.
    """
    final_prompt = ""
    try:
        # --- Replicating the logic from the non-SDK mode ---
        content_template = config_dict.get("content_template", "")
        formatted_content = content_template
        for col_name, cell_value in record_data.items():
            placeholder = f"{{row['{col_name}']}}"
            value_str = str(cell_value) if pd.notna(cell_value) else ""
            formatted_content = formatted_content.replace(placeholder, value_str)

        llm_template = config_dict.get("llm_template", "")
        final_prompt = llm_template.replace("{{content}}", formatted_content)

        output_columns = config_dict.get("output_columns", [])
        if output_columns:
            output_format_prompt = ", ".join([f'\"{col}\": \"...\"' for col in output_columns])
            final_prompt += f"\n\nPlease provide the output in a single, valid JSON object format, like this: {{{output_format_prompt}}}. Do not include any text or formatting outside of the JSON object."

        sdk_record = {
            "model": config_dict.get("model"),
            "messages": [{"role": "user", "content": final_prompt}],
            "extra_headers": {},
        }

        result = await client.batch_chat.completions.create(**sdk_record)
        result_dict = result.to_dict()

        # CORRECTED LOGIC: Extract and parse the 'content' string
        content_str = result_dict.get("choices", [{}])[0].get("message", {}).get("content", "")
        if not content_str:
            raise Exception(f"LLM response content is empty. Full API response: {result_dict}")

        try:
            start_index = content_str.find('{')
            end_index = content_str.rfind('}') + 1
            if start_index == -1 or end_index == 0:
                raise json.JSONDecodeError("No JSON object found in response content", content_str, 0)
            
            json_str = content_str[start_index:end_index]
            parsed_data = json.loads(json_str, strict=False)

            if not isinstance(parsed_data, dict):
                raise TypeError(f"Parsed data is not a dictionary. Got: {type(parsed_data)}")

        except (json.JSONDecodeError, TypeError) as e:
            raise Exception(f"Failed to parse LLM content. Error: {e}. Raw content: '{content_str}'")

        final_record = {**record_data, **parsed_data}
        f_out.write(json.dumps(final_record, ensure_ascii=False) + '\n')
        progress_queue.put(1)

    except Exception as e:
        error_info = {
            **record_data, 
            "__error__": str(e),
            "__prompt_sent__": final_prompt
        }
        f_out.write(json.dumps(error_info, ensure_ascii=False) + '\n')
        progress_queue.put(1)
    finally:
        sem.release()


class VolcengineProcessor:
    def __init__(self, config: ProcessingConfig, progress_queue: multiprocessing.Queue):
        self.config = config
        self.progress_queue = progress_queue
        self.api_key = self.config.api_key or os.environ.get("ARK_API_KEY")
        self.num_worker_processes = self.config.workers or (os.cpu_count() or 1)
        self.max_concurrency_per_process = 64
        self.temp_dir = tempfile.mkdtemp(prefix="excel_proc_")
        self.input_jsonl_path = os.path.join(self.temp_dir, "input.jsonl")
        self.output_paths = [os.path.join(self.temp_dir, f"output_{i}.jsonl") for i in range(self.num_worker_processes)]

    def _prepare_input_file(self) -> int:
        try:
            workbook = openpyxl.load_workbook(self.config.input_file, read_only=True)
            sheet = workbook[self.config.sheet_name]
            
            total_rows = 0
            header = [cell.value for cell in sheet[1]]
            empty_col_index = header.index(self.config.empty_column) if self.config.empty_column in header else -1

            with open(self.input_jsonl_path, 'w', encoding='utf-8') as f:
                # iter_rows(min_row=2) skips the header row
                for row_cells in sheet.iter_rows(min_row=2):
                    # Check if the row should be skipped
                    if empty_col_index != -1 and (row_cells[empty_col_index].value is None or str(row_cells[empty_col_index].value).strip() == ""):
                        continue

                    row_data = {header[i]: cell.value for i, cell in enumerate(row_cells)}
                    f.write(json.dumps(row_data, ensure_ascii=False) + '\n')
                    total_rows += 1
            return total_rows
        except Exception as e:
            raise RuntimeError(f"Failed to prepare input file using openpyxl: {e}")

    def run(self):
        try:
            total_rows = self._prepare_input_file()
            self.progress_queue.put(("total_rows", total_rows, 0))
            if total_rows == 0:
                return

            processes = []
            config_dict = dataclasses.asdict(self.config)

            for i in range(self.num_worker_processes):
                p = multiprocessing.Process(
                    target=process_entrypoint,
                    args=(
                        i, self.num_worker_processes, self.max_concurrency_per_process,
                        self.api_key, self.input_jsonl_path, self.output_paths[i],
                        config_dict, self.progress_queue
                    ),
                    daemon=True
                )
                p.start()
                processes.append(p)

            for p in processes:
                p.join()
            
            self._merge_and_save_results()

        except Exception as e:
            self.progress_queue.put(("error", f"处理失败: {e}", 0))
        finally:
            self._cleanup_temp_files()

    def _merge_and_save_results(self):
        all_results = []
        for path in self.output_paths:
            if os.path.exists(path):
                # Open with error handling for robustness, in case any worker still writes bad data.
                with open(path, 'r', encoding='utf-8', errors='replace') as f:
                    for line in f:
                        try:
                            all_results.append(json.loads(line))
                        except json.JSONDecodeError:
                            print(f"Warning: Could not decode line in {path}: {line}")
                            continue
        
        if not all_results:
            self.progress_queue.put(("info", "没有生成任何结果。", 0))
            return

        self.progress_queue.put(("info", "正在合并结果并生成最终Excel文件...", 0))
        final_df = pd.DataFrame(all_results)
        
        # Ensure output directory exists
        output_dir = os.path.dirname(self.config.output_file)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)
            
        final_df.to_excel(self.config.output_file, index=False)

    def _cleanup_temp_files(self):
        try:
            if os.path.exists(self.temp_dir):
                import shutil
                shutil.rmtree(self.temp_dir)
        except Exception as e:
            print(f"Failed to clean up temp directory {self.temp_dir}: {e}")
