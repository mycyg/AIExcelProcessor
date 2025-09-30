import pandas as pd
import traceback
import os
import json
import requests
import time
import tempfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Iterator, Dict, Any, Tuple, List

from config import ProcessingConfig

class ExcelProcessor:
    def __init__(self, config: ProcessingConfig):
        self.config = config
        self.should_stop = False
        self.total_rows = 0
        self.temp_files = []
        self.session = requests.Session()
        self.sheet = None
        self.headers = []

    def stop(self):
        self.should_stop = True

    def _prepare_and_count_rows(self):
        """
        Opens the Excel file, reads the headers, and performs an initial pass
        to get an accurate count of rows that will actually be processed,
        respecting the 'empty_column' filter.
        """
        from openpyxl import load_workbook
        try:
            workbook = load_workbook(filename=self.config.input_file, read_only=True)
            self.sheet = workbook[self.config.sheet_name]
            self.headers = [cell.value for cell in self.sheet[1]]
            
            if not self.headers:
                self.total_rows = 0
                return

            empty_col_index = -1
            if self.config.empty_column and self.config.empty_column in self.headers:
                empty_col_index = self.headers.index(self.config.empty_column)

            count = 0
            for row in self.sheet.iter_rows(min_row=2, values_only=True):
                if empty_col_index != -1:
                    # Check if the cell in the empty_column is None or an empty string
                    if row[empty_col_index] is None or str(row[empty_col_index]).strip() == "":
                        continue
                count += 1
            self.total_rows = count

        except Exception as e:
            # Re-raise exceptions to be caught in start_processing
            raise RuntimeError(f"准备读取Excel文件并计数时出错: {e}") from e

    def start_processing(self) -> Iterator[Tuple[str, Any, Any]]:
        self.should_stop = False
        self.temp_files = []
        processed_count = 0

        try:
            yield "info", "正在准备并计算有效行数...", 0
            self._prepare_and_count_rows()
            
            yield "info", f"预计总行数: {self.total_rows}", 0
            if self.total_rows == 0:
                yield "finish", 0, 0
                return

            yield "info", "正在以流式模式读取数据...", 0
            row_iterator = self._read_and_filter_data()

            with ThreadPoolExecutor(max_workers=self.config.workers) as executor:
                futures = {}
                batch_num = 0
                
                while not self.should_stop:
                    batch = self._get_next_batch_from_iterator(row_iterator)
                    if not batch:
                        break
                    
                    future = executor.submit(self._process_batch, batch)
                    futures[future] = batch_num
                    batch_num += 1

                for future in as_completed(futures):
                    if self.should_stop:
                        break
                    
                    for result_type, data, total in future.result():
                        if result_type == "data":
                            temp_file_path, batch_row_count = data
                            self.temp_files.append(temp_file_path)
                            processed_count += batch_row_count
                            yield "progress", processed_count, self.total_rows
                        else:
                            yield result_type, data, total
            
            if self.should_stop:
                yield "stopped", processed_count, self.total_rows
                return

            if self.temp_files:
                yield "info", "正在合并临时文件...", 0
                self._merge_temp_files()
                yield "finish", processed_count, self.total_rows
            else:
                # This case can happen if all rows were processed but resulted in errors
                yield "finish", processed_count, self.total_rows

        except Exception as e:
            yield "error", f"处理过程中发生未知错误: {e}", 0
            traceback.print_exc()
        finally:
            self._cleanup_temp_files()

    def _get_next_batch_from_iterator(self, iterator: Iterator[Dict[str, Any]]) -> List[Dict[str, Any]]:
        batch = []
        try:
            for _ in range(self.config.batch_size):
                batch.append(next(iterator))
        except StopIteration:
            pass
        return batch

    def _read_and_filter_data(self) -> Iterator[Dict[str, Any]]:
        """
        A generator that yields rows from the pre-loaded sheet, applying the
        'empty_column' filter again to ensure it yields the same rows that
        were counted.
        """
        try:
            empty_col_index = -1
            if self.config.empty_column and self.config.empty_column in self.headers:
                empty_col_index = self.headers.index(self.config.empty_column)

            for row in self.sheet.iter_rows(min_row=2, values_only=True):
                if self.should_stop:
                    break
                
                if empty_col_index != -1:
                    if row[empty_col_index] is None or str(row[empty_col_index]).strip() == "":
                        continue
                
                row_data = dict(zip(self.headers, row))
                yield row_data
        except Exception as e:
            # This is a generator, so we can't easily yield an error message from here.
            # The error will be raised when the generator is consumed in start_processing.
            print(f"Error reading excel file row-by-row: {e}")
            traceback.print_exc()
            raise

    def _process_batch(self, batch: List[Dict[str, Any]]) -> List[Tuple[str, Any, Any]]:
        batch_results = []
        log_results = []
        for row in batch:
            if self.should_stop:
                break
            try:
                content = self._format_content(row)
                log_results.append(("debug_prompt", content, 0))
                api_response = self._call_api(content)
                log_results.append(("debug_response", api_response, 0))
                parsed_result = self._parse_llm_response(api_response)
                
                final_row = {col: row.get(col) for col, selected in self.config.input_columns.items() if selected}
                final_row.update(parsed_result)
                batch_results.append(final_row)
            except Exception as e:
                log_results.append(("error", f"处理行失败: {e}", 0))
                traceback.print_exc()
                continue
        
        if not batch_results:
            return log_results

        try:
            temp_fd, temp_path = tempfile.mkstemp(suffix=".xlsx", prefix="proc_")
            os.close(temp_fd)
            result_df = pd.DataFrame(batch_results)
            result_df.to_excel(temp_path, index=False)
            log_results.append(("data", (temp_path, len(batch_results)), 0))
            return log_results
        except Exception as e:
            log_results.append(("error", f"保存临时文件失败: {e}", 0))
            traceback.print_exc()
            return log_results

    def _format_content(self, row: Dict[str, Any]) -> str:
        formatted_content = self.config.content_template
        for col_name, cell_value in row.items():
            placeholder = f"{{row['{col_name}']}}"
            value_str = str(cell_value) if cell_value is not None else ""
            formatted_content = formatted_content.replace(placeholder, value_str)
        return formatted_content

    def _call_api(self, content: str) -> str:
        output_format_prompt = ", ".join([f'\"{col}\": \"...\"' for col in self.config.output_columns])
        prompt = self.config.llm_template.replace('{{content}}', content)
        prompt += f"\n\nPlease provide the output in a single, valid JSON object format, like this: {{{output_format_prompt}}}. Do not include any text or formatting outside of the JSON object."

        headers = {
            'Authorization': f'Bearer {self.config.api_key}',
            'Content-Type': 'application/json'
        }
        data = {
            'model': self.config.model,
            'messages': [{'role': 'user', 'content': prompt}]
        }
        
        max_retries = 3
        retry_delay = 1
        for attempt in range(max_retries):
            try:
                response = self.session.post(
                    self.config.api_url, 
                    headers=headers, 
                    json=data, 
                    timeout=self.config.api_timeout
                )
                response.raise_for_status()
                return response.json()['choices'][0]['message']['content']
            except requests.exceptions.RequestException as e:
                print(f"API call failed (attempt {attempt + 1}/{max_retries}): {e}")
                if attempt < max_retries - 1:
                    time.sleep(retry_delay)
                else:
                    raise

    def _parse_llm_response(self, response: str) -> Dict[str, Any]:
        try:
            start_index = response.find('{')
            end_index = response.rfind('}') + 1
            if start_index == -1 or end_index == 0:
                raise json.JSONDecodeError("No JSON object found in response", response, 0)
            json_str = response[start_index:end_index]
            return json.loads(json_str, strict=False)
        except json.JSONDecodeError as e:
            print(f"Failed to parse LLM response as JSON: {e}")
            return {col: "PARSE_ERROR" for col in self.config.output_columns}

    def _merge_temp_files(self):
        from openpyxl import load_workbook, Workbook

        if not self.temp_files:
            return

        # Create a new workbook for the final output
        final_workbook = Workbook()
        final_sheet = final_workbook.active

        # Use the first temp file to write the header
        first_file = self.temp_files[0]
        try:
            source_workbook = load_workbook(filename=first_file, read_only=True)
            source_sheet = source_workbook.active
            
            # Write header
            header = [cell.value for cell in source_sheet[1]]
            final_sheet.append(header)
            
            # Write data rows from the first file
            for row in source_sheet.iter_rows(min_row=2, values_only=True):
                final_sheet.append(row)
            
            source_workbook.close()

            # Process remaining temp files
            for temp_file in self.temp_files[1:]:
                try:
                    source_workbook = load_workbook(filename=temp_file, read_only=True)
                    source_sheet = source_workbook.active
                    # Append data rows, skipping the header
                    for row in source_sheet.iter_rows(min_row=2, values_only=True):
                        final_sheet.append(row)
                    source_workbook.close()
                except Exception as e:
                    print(f"Error processing temp file {temp_file}: {e}")

            output_dir = os.path.dirname(self.config.output_file)
            if output_dir:
                os.makedirs(output_dir, exist_ok=True)
            final_workbook.save(self.config.output_file)

        except Exception as e:
            print(f"Error merging temp files: {e}")
            traceback.print_exc()

    def _cleanup_temp_files(self):
        for f in self.temp_files:
            try:
                os.remove(f)
            except OSError as e:
                print(f"Error deleting temp file {f}: {e}")
        self.temp_files = []
